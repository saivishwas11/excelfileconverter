import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
import math
import re
import webcolors
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import io
import logging
from datetime import datetime
import os

st.set_page_config(page_title="HTML to Excel Converter", layout="centered")

hide_github_icon = """
    <style>
    a[aria-label="View source code"] {
        display: none !important;
    }
    </style>
"""
st.markdown(hide_github_icon, unsafe_allow_html=True)
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ALLOWED_EXTENSIONS = {'html', 'htm'}
MAX_FILE_SIZE_MB = 200
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def html_color_to_openpyxl_argb(html_color):
    if not html_color:
        return None
    html_color = html_color.lower().strip()
    try:
        if html_color.startswith('#'):
            hex_val = html_color.lstrip('#')
        else:
            hex_val = webcolors.name_to_hex(html_color).lstrip('#')
        if len(hex_val) == 3:
            hex_val = "".join([c*2 for c in hex_val])
        if len(hex_val) == 6:
            return 'FF' + hex_val.upper()
        else:
            return None
    except ValueError:
        return None

def convert_to_excel(input_html, output_stream):
    PIXELS_TO_EXCEL_UNITS = 8.43
    soup = BeautifulSoup(input_html, 'html.parser')
    tables = soup.find_all('table')
    if not tables:
        text = soup.get_text(separator='\n', strip=True)
        df = pd.DataFrame([line for line in text.split('\n') if line], columns=['Content'])
        df.to_excel(output_stream, index=False)
        return
    workbook = Workbook()
    worksheet = workbook.active
    thin_black_side = Side(style='thin', color='FF000000')
    default_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)
    master_layout_pixels = []
    max_cols = 0
    for table in tables:
        cols = table.find_all('col')
        if len(cols) > max_cols:
            max_cols = len(cols)
            master_layout_pixels = []
            for col in cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match:
                    master_layout_pixels.append(int(match.group(1)))
    if not master_layout_pixels:
        logger.error("Could not determine a master layout from <colgroup> tags.")
        pd.read_html(str(soup)).to_excel(output_stream, index=False)
        return
    master_layout_excel_units = [px / PIXELS_TO_EXCEL_UNITS for px in master_layout_pixels]
    for i, width in enumerate(master_layout_excel_units):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = width
    current_row_excel = 1
    for table in tables:
        local_layout_pixels = []
        local_cols = table.find_all('col')
        if local_cols:
            for col in local_cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match: local_layout_pixels.append(int(match.group(1)))
        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            current_col_excel = 1
            for cell_idx, cell in enumerate(cells):
                text = cell.get_text(strip=True)
                style_str = cell.get('style', '') + row.get('style', '')
                bg_color_html = cell.get('bgcolor')
                if not bg_color_html:
                    bg_match = re.search(r'background-color:\s*([^;]+)', style_str)
                    if bg_match: bg_color_html = bg_match.group(1).strip()
                font_color_html = None
                color_match = re.search(r'(?<!background-)color:\s*([^;]+)', style_str)
                if color_match: font_color_html = color_match.group(1).strip()
                align_map = {'center': 'center', 'left': 'left', 'right': 'right', 'justify': 'justify'}
                text_align = 'general'
                align_match = re.search(r'text-align:\s*([^;]+)', style_str)
                if align_match: text_align = align_map.get(align_match.group(1).strip().lower(), 'general')
                is_bold = 'font-weight: bold' in style_str or cell.find('b') or cell.name == 'th'
                font_family = None
                font_size = None
                is_italic = 'font-style: italic' in style_str or cell.find('i')
                is_underline = 'text-decoration: underline' in style_str
                is_strike = 'text-decoration: line-through' in style_str
                font_family_match = re.search(r'font-family:\s*([^;]+)', style_str)
                if font_family_match:
                    font_family = font_family_match.group(1).split(',')[0].strip().strip("'\"")
                font_size_match = re.search(r'font-size:\s*([\d.]+)px', style_str)
                if font_size_match:
                    font_size = float(font_size_match.group(1)) / 1.33
                html_colspan = int(cell.get('colspan', 1))
                target_pixel_width = 0
                if local_layout_pixels and cell_idx < len(local_layout_pixels):
                    for i in range(html_colspan):
                        if (cell_idx + i) < len(local_layout_pixels):
                            target_pixel_width += local_layout_pixels[cell_idx + i]
                excel_colspan = 0
                covered_width = 0
                if target_pixel_width > 0:
                    start_master_col_idx = current_col_excel - 1
                    while covered_width < (target_pixel_width * 0.9) and (start_master_col_idx + excel_colspan) < len(master_layout_pixels):
                        covered_width += master_layout_pixels[start_master_col_idx + excel_colspan]
                        excel_colspan += 1
                excel_colspan = max(1, excel_colspan)
                alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)
                font = Font(
                    name=font_family if font_family else None,
                    size=font_size if font_size else None,
                    bold=bool(is_bold),
                    italic=bool(is_italic),
                    underline='single' if is_underline else None,
                    strike=bool(is_strike),
                    color=html_color_to_openpyxl_argb(font_color_html)
                )
                fill = None
                bg_color_argb = html_color_to_openpyxl_argb(bg_color_html)
                if bg_color_argb:
                    try: fill = PatternFill(start_color=bg_color_argb, end_color=bg_color_argb, fill_type="solid")
                    except ValueError: fill = None
                target_cell = worksheet.cell(row=current_row_excel, column=current_col_excel)
                target_cell.value = text
                target_cell.alignment = alignment
                if fill: target_cell.fill = fill
                target_cell.font = font
                if excel_colspan > 1:
                    end_col = current_col_excel + excel_colspan - 1
                    worksheet.merge_cells(start_row=current_row_excel, start_column=current_col_excel, end_row=current_row_excel, end_column=end_col)
                    for r_offset in range(1):
                        for c_offset in range(excel_colspan):
                             worksheet.cell(row=current_row_excel + r_offset, column=current_col_excel + c_offset).border = default_border
                else:
                    target_cell.border = default_border
                current_col_excel += excel_colspan
            current_row_excel += 1
        current_row_excel += 1
    POINTS_PER_LINE = 15.0
    for row_index in range(1, worksheet.max_row + 1):
        max_lines_in_row = 1
        for cell in worksheet[row_index]:
            if not cell.value: continue
            effective_width_units = 0
            is_merged = False
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                        effective_width_units += worksheet.column_dimensions[get_column_letter(col_idx)].width
                    is_merged = True
                    break
            if not is_merged:
                effective_width_units = worksheet.column_dimensions[cell.column_letter].width
            text = str(cell.value)
            lines_from_newlines = text.count('\n') + 1
            lines_from_wrapping = 1
            if effective_width_units > 0:
                lines_from_wrapping = math.ceil(len(text) / (effective_width_units / 1.1))
            cell_lines = max(lines_from_newlines, lines_from_wrapping)
            if cell_lines > max_lines_in_row:
                max_lines_in_row = cell_lines
        worksheet.row_dimensions[row_index].height = max_lines_in_row * POINTS_PER_LINE
    workbook.save(output_stream)
    output_stream.seek(0)

# --- UI Layout ---
st.markdown("""
<div style='text-align: center;'>
    <h1 style='font-size:2.5rem;'>📄➡️📊 HTML to Excel Converter</h1>
    <p style='font-size:1.1rem;'>Easily convert your HTML or HTM files to Excel (.xlsx) format.<br>
    <b>Limit 200MB per file • HTML, HTM</b></p>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<ol style='font-size:1.05rem;'>
  <li>Upload your <b>HTML</b> or <b>HTM</b> file below.</li>
  <li>Click <b>Convert</b> to instantly download your Excel file.</li>
</ol>
""", unsafe_allow_html=True)

col1, col2, col3 = st.columns([1,2,1])
with col2:
    uploaded_file = st.file_uploader("", type=["html", "htm"], label_visibility="collapsed")
    if uploaded_file is not None:
        if not allowed_file(uploaded_file.name):
            st.error(f"❌ Unsupported file type. Allowed types: {', '.join([ext.upper() for ext in ALLOWED_EXTENSIONS])}")
        elif uploaded_file.size > MAX_FILE_SIZE_BYTES:
            st.error(f"❌ File size exceeds {MAX_FILE_SIZE_MB}MB limit.")
        else:
            file_base = os.path.splitext(uploaded_file.name)[0]
            excel_filename = f"{file_base}.xlsx"
            convert_clicked = st.button("🚀 Convert & Download", use_container_width=True)
            if convert_clicked:
                try:
                    html_content = uploaded_file.read().decode('utf-8')
                    output_stream = io.BytesIO()
                    convert_to_excel(html_content, output_stream)
                    st.success("✅ Conversion successful! Your download should begin below.")
                    st.download_button(
                        label="⬇️ Download Excel file",
                        data=output_stream,
                        file_name=excel_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Error during conversion: {str(e)}")