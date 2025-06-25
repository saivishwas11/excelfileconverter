from flask import Flask, request, send_file, abort, render_template, jsonify
import os
import tempfile
from werkzeug.utils import secure_filename
import pandas as pd
from bs4 import BeautifulSoup
import math
import re
import webcolors
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import logging
import traceback
import uuid
import magic
from flask_cors import CORS
import base64
from datetime import datetime

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

app = Flask(__name__, template_folder='templates')


CORS(app, resources={
    r"/api/*": {
        "origins": ["*"],  # You should restrict this to your Appian domain in production
        "methods": ["POST", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"],
        "max_age": 3600
    }
})

ALLOWED_EXTENSIONS = {'html'}
MIME_TYPES = {
    'html': 'text/html'
}
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB

@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Access-Control-Allow-Origin'] = '*'  # Restrict this in production
    response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response

# Health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'version': '1.0.0'
    })

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validate_mime_type(filepath, ext):
    try:
        mime = magic.from_file(filepath, mime=True)
        expected_mime = MIME_TYPES.get(ext)
        logger.debug(f"File MIME type: {mime}, Expected MIME type: {expected_mime}")
        return mime == expected_mime
    except Exception as e:
        logger.error(f"Error validating MIME type for {filepath}: {e}")
        return False

def convert_to_excel(input_file, output_file):
    PIXELS_TO_EXCEL_UNITS = 8.43 

    with open(input_file, 'r', encoding='utf-8') as f:
        html_content = f.read()

    soup = BeautifulSoup(html_content, 'html.parser')
    tables = soup.find_all('table')

    if not tables:
        text = soup.get_text(separator='\n', strip=True)
        df = pd.DataFrame([line for line in text.split('\n') if line], columns=['Content'])
        df.to_excel(output_file, index=False)
        return

    workbook = Workbook()
    worksheet = workbook.active
    
    thin_black_side = Side(style='thin', color='FF000000')
    default_border = Border(left=thin_black_side, right=thin_black_side, top=thin_black_side, bottom=thin_black_side)

    master_layout_pixels = []
    max_cols = 0
    for table in tables:
        cols = table.find_all('col')
        if len(cols) > max_cols:
            max_cols = len(cols)
            master_layout_pixels = []
            for col in cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match:
                    master_layout_pixels.append(int(match.group(1)))

    if not master_layout_pixels:
        logger.error("Could not determine a master layout from <colgroup> tags.")
        pd.read_html(html_content).to_excel(output_file, index=False)
        return

    master_layout_excel_units = [px / PIXELS_TO_EXCEL_UNITS for px in master_layout_pixels]
    for i, width in enumerate(master_layout_excel_units):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = width

    current_row_excel = 1
    for table in tables:
        local_layout_pixels = []
        local_cols = table.find_all('col')
        if local_cols:
            for col in local_cols:
                style = col.get('style', '')
                match = re.search(r'width:\s*(\d+)', style)
                if match: local_layout_pixels.append(int(match.group(1)))

        rows = table.find_all('tr')
        for row in rows:
            cells = row.find_all(['td', 'th'])
            current_col_excel = 1
            
            for cell_idx, cell in enumerate(cells):
                text = cell.get_text(strip=True)
                style_str = cell.get('style', '') + row.get('style', '')
                
                bg_color_html = cell.get('bgcolor')
                if not bg_color_html:
                    bg_match = re.search(r'background-color:\s*([^;]+)', style_str)
                    if bg_match: bg_color_html = bg_match.group(1).strip()
                font_color_html = None
                color_match = re.search(r'(?<!background-)color:\s*([^;]+)', style_str)
                if color_match: font_color_html = color_match.group(1).strip()
                align_map = {'center': 'center', 'left': 'left', 'right': 'right', 'justify': 'justify'}
                text_align = 'general'
                align_match = re.search(r'text-align:\s*([^;]+)', style_str)
                if align_match: text_align = align_map.get(align_match.group(1).strip().lower(), 'general')
                is_bold = 'font-weight: bold' in style_str or cell.find('b') or cell.name == 'th'

                # Extract font properties
                font_family = None
                font_size = None
                is_italic = 'font-style: italic' in style_str or cell.find('i')
                is_underline = 'text-decoration: underline' in style_str
                is_strike = 'text-decoration: line-through' in style_str

                # Regex for font-family and font-size
                font_family_match = re.search(r'font-family:\s*([^;]+)', style_str)
                if font_family_match:
                    font_family = font_family_match.group(1).split(',')[0].strip().strip("'\"")

                font_size_match = re.search(r'font-size:\s*([\d.]+)px', style_str)
                if font_size_match:
                    # Convert px to points (1pt ≈ 1.33px)
                    font_size = float(font_size_match.group(1)) / 1.33

                html_colspan = int(cell.get('colspan', 1))
                
                target_pixel_width = 0
                if local_layout_pixels and cell_idx < len(local_layout_pixels):
                    for i in range(html_colspan):
                        if (cell_idx + i) < len(local_layout_pixels):
                            target_pixel_width += local_layout_pixels[cell_idx + i]
                
                excel_colspan = 0
                covered_width = 0
                if target_pixel_width > 0:
                    start_master_col_idx = current_col_excel - 1
                    while covered_width < (target_pixel_width * 0.9) and (start_master_col_idx + excel_colspan) < len(master_layout_pixels):
                        covered_width += master_layout_pixels[start_master_col_idx + excel_colspan]
                        excel_colspan += 1
                excel_colspan = max(1, excel_colspan)

                alignment = Alignment(horizontal=text_align, vertical='center', wrap_text=True)
                font = Font(
                    name=font_family if font_family else None,
                    size=font_size if font_size else None,
                    bold=bool(is_bold),
                    italic=bool(is_italic),
                    underline='single' if is_underline else None,
                    strike=bool(is_strike),
                    color=html_color_to_openpyxl_argb(font_color_html)
                )
                fill = None
                bg_color_argb = html_color_to_openpyxl_argb(bg_color_html)
                if bg_color_argb:
                    try: fill = PatternFill(start_color=bg_color_argb, end_color=bg_color_argb, fill_type="solid")
                    except ValueError: fill = None

                target_cell = worksheet.cell(row=current_row_excel, column=current_col_excel)
                target_cell.value = text
                target_cell.alignment = alignment
                if fill: target_cell.fill = fill
                target_cell.font = font

                if excel_colspan > 1:
                    end_col = current_col_excel + excel_colspan - 1
                    worksheet.merge_cells(start_row=current_row_excel, start_column=current_col_excel, end_row=current_row_excel, end_column=end_col)
                    for r_offset in range(1):
                        for c_offset in range(excel_colspan):
                             worksheet.cell(row=current_row_excel + r_offset, column=current_col_excel + c_offset).border = default_border
                else:
                    target_cell.border = default_border
                
                current_col_excel += excel_colspan
            current_row_excel += 1
        current_row_excel += 1

    POINTS_PER_LINE = 15.0
    for row_index in range(1, worksheet.max_row + 1):
        max_lines_in_row = 1
        for cell in worksheet[row_index]:
            if not cell.value: continue
            
            effective_width_units = 0
            is_merged = False
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                        effective_width_units += worksheet.column_dimensions[get_column_letter(col_idx)].width
                    is_merged = True
                    break
            if not is_merged:
                effective_width_units = worksheet.column_dimensions[cell.column_letter].width

            text = str(cell.value)
            lines_from_newlines = text.count('\n') + 1
            lines_from_wrapping = 1
            if effective_width_units > 0:
                lines_from_wrapping = math.ceil(len(text) / (effective_width_units / 1.1))

            cell_lines = max(lines_from_newlines, lines_from_wrapping)
            if cell_lines > max_lines_in_row:
                max_lines_in_row = cell_lines
        
        worksheet.row_dimensions[row_index].height = max_lines_in_row * POINTS_PER_LINE

    workbook.save(output_file)
    
def html_color_to_openpyxl_argb(html_color):
    if not html_color:
        return None
    
    html_color = html_color.lower().strip()
    
    try:
        if html_color.startswith('#'):
            hex_val = html_color.lstrip('#')
        else:
            hex_val = webcolors.name_to_hex(html_color).lstrip('#')

        if len(hex_val) == 3:
            hex_val = "".join([c*2 for c in hex_val])
        
        if len(hex_val) == 6:
            return 'FF' + hex_val.upper()
        else:
            return None
            
    except ValueError:
        return None
        
@app.route('/api/convert', methods=['POST'])
def convert_html_to_excel():
    """
    API endpoint to convert HTML to Excel
    Expected JSON payload:
    {
        "html_content": "base64_encoded_html_content"
    }
    """
    try:
        if not request.is_json:
            return jsonify({
                'error': 'Content-Type must be application/json'
            }), 400

        data = request.get_json()
        if not data or 'html_content' not in data:
            return jsonify({
                'error': 'Missing html_content in request body'
            }), 400

        # Validate base64 content
        html_content_b64 = data['html_content']
        if not isinstance(html_content_b64, str):
            return jsonify({
                'error': 'html_content must be a string'
            }), 400

        # Check if the base64 string is valid
        try:
            # Try to decode a small portion first to validate format
            base64.b64decode(html_content_b64[:100])
        except Exception:
            return jsonify({
                'error': 'Invalid base64 format'
            }), 400

        # Decode full content
        try:
            html_content = base64.b64decode(html_content_b64).decode('utf-8')
        except UnicodeDecodeError:
            return jsonify({
                'error': 'Invalid UTF-8 encoding in HTML content'
            }), 400
        except Exception as e:
            return jsonify({
                'error': 'Error decoding base64 content',
                'details': str(e)
            }), 400

        # Validate HTML content
        if not html_content.strip():
            return jsonify({
                'error': 'Empty HTML content'
            }), 400

        with tempfile.TemporaryDirectory() as tmpdirname:
            # Save HTML content to temporary file
            input_file = os.path.join(tmpdirname, 'input.html')
            with open(input_file, 'w', encoding='utf-8') as f:
                f.write(html_content)

            # Convert to Excel
            output_file = os.path.join(tmpdirname, 'converted.xlsx')
            try:
                convert_to_excel(input_file, output_file)
            except Exception as e:
                logger.error(f"Error during Excel conversion: {str(e)}")
                return jsonify({
                    'error': 'Error converting HTML to Excel',
                    'details': str(e)
                }), 500

            # Read the Excel file and convert to base64
            try:
                with open(output_file, 'rb') as f:
                    excel_content = f.read()
                    excel_base64 = base64.b64encode(excel_content).decode('utf-8')
            except Exception as e:
                logger.error(f"Error reading Excel file: {str(e)}")
                return jsonify({
                    'error': 'Error processing Excel file',
                    'details': str(e)
                }), 500

            return jsonify({
                'success': True,
                'excel_content': excel_base64,
                'filename': 'converted.xlsx',
                'timestamp': datetime.utcnow().isoformat()
            })

    except Exception as e:
        logger.error(f"Error during conversion: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Internal server error during conversion',
            'details': str(e)
        }), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    temp_output = None
    output_extension = '.xlsx'

    try:
        with tempfile.TemporaryDirectory() as tmpdirname:
            filepath = None
            output_file = None

            if 'file' not in request.files:
                logger.error("No file part in the request")
                abort(400, 'No file part in the request.')
            
            file = request.files['file']
            if file.filename == '':
                logger.error("No selected file")
                abort(400, 'No selected file.')
            
            if not allowed_file(file.filename):
                logger.error(f"Unsupported file type: {file.filename}")
                abort(400, f'Unsupported file type. Allowed types: {", ".join(ALLOWED_EXTENSIONS)}')

            filename = secure_filename(file.filename)
            filepath = os.path.join(tmpdirname, filename)
            file.save(filepath)
            ext = filename.rsplit('.', 1)[1].lower()

            if not validate_mime_type(filepath, ext):
                logger.error(f"File type mismatch for {filename}")
                abort(400, 'File type mismatch. Possible malicious or corrupted file.')

            output_file = os.path.join(tmpdirname, f'converted{output_extension}')

            try:
                convert_to_excel(filepath, output_file)

                temp_output = os.path.join(tempfile.gettempdir(), f'converted_{uuid.uuid4().hex}{output_extension}')
                with open(output_file, 'rb') as src, open(temp_output, 'wb') as dst:
                    dst.write(src.read())

            except Exception as e:
                logger.error(f"Error during file conversion: {str(e)}")
                logger.error(traceback.format_exc())
                abort(500, f'Error during file conversion: {str(e)}')

    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        logger.error(traceback.format_exc())
        abort(500, 'Internal server error.')

    if not temp_output or not os.path.exists(temp_output):
        logger.error("Output file was not created or found")
        abort(500, 'Failed to create the output file.')

    download_filename = f'converted_file{output_extension}'
    return send_file(temp_output, as_attachment=True, download_name=download_filename)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)